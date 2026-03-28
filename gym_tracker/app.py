from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import sqlite3
from datetime import datetime, timedelta
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DB_PATH = "gym_tracker.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS workouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            exercise TEXT NOT NULL,
            sets INTEGER,
            reps INTEGER,
            weight_kg REAL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS body_stats (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            weight_kg REAL,
            body_fat_pct REAL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS goals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,         -- 'strength' or 'bodyweight'
            exercise TEXT,              -- for strength goals
            target_kg REAL NOT NULL,
            start_kg REAL,
            deadline TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()

    # Seed with some example data if empty
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM workouts")
    if cur.fetchone()[0] == 0:
        today = datetime.now()
        seed_workouts = [
            ((today - timedelta(days=14)).strftime("%Y-%m-%d"), "Bench Press", 4, 8, 80, "Felt strong"),
            ((today - timedelta(days=12)).strftime("%Y-%m-%d"), "Squat", 4, 6, 100, ""),
            ((today - timedelta(days=12)).strftime("%Y-%m-%d"), "Deadlift", 3, 5, 120, "New PR"),
            ((today - timedelta(days=9)).strftime("%Y-%m-%d"), "Bench Press", 4, 8, 82.5, ""),
            ((today - timedelta(days=7)).strftime("%Y-%m-%d"), "Pull-up", 3, 10, 0, "Bodyweight"),
            ((today - timedelta(days=7)).strftime("%Y-%m-%d"), "Squat", 4, 6, 102.5, ""),
            ((today - timedelta(days=5)).strftime("%Y-%m-%d"), "Overhead Press", 4, 8, 55, ""),
            ((today - timedelta(days=3)).strftime("%Y-%m-%d"), "Deadlift", 3, 5, 125, "PR!"),
            ((today - timedelta(days=3)).strftime("%Y-%m-%d"), "Bench Press", 4, 9, 82.5, "Extra rep"),
            ((today - timedelta(days=1)).strftime("%Y-%m-%d"), "Squat", 4, 7, 105, ""),
        ]
        conn.executemany(
            "INSERT INTO workouts (date, exercise, sets, reps, weight_kg, notes) VALUES (?,?,?,?,?,?)",
            seed_workouts
        )

        seed_stats = [
            ((today - timedelta(days=14)).strftime("%Y-%m-%d"), 82.5, 18.0, ""),
            ((today - timedelta(days=10)).strftime("%Y-%m-%d"), 82.0, 17.8, ""),
            ((today - timedelta(days=7)).strftime("%Y-%m-%d"), 81.5, 17.5, ""),
            ((today - timedelta(days=3)).strftime("%Y-%m-%d"), 81.0, 17.2, ""),
            (today.strftime("%Y-%m-%d"), 80.5, 17.0, "Feeling leaner"),
        ]
        conn.executemany(
            "INSERT INTO body_stats (date, weight_kg, body_fat_pct, notes) VALUES (?,?,?,?)",
            seed_stats
        )

        seed_goals = [
            ("strength", "Squat",       140.0, 100.0, (today + timedelta(days=60)).strftime("%Y-%m-%d")),
            ("strength", "Bench Press", 100.0,  80.0, (today + timedelta(days=90)).strftime("%Y-%m-%d")),
            ("bodyweight", None,         77.0,  82.5, (today + timedelta(days=45)).strftime("%Y-%m-%d")),
        ]
        conn.executemany(
            "INSERT INTO goals (type, exercise, target_kg, start_kg, deadline) VALUES (?,?,?,?,?)",
            seed_goals
        )
        conn.commit()
    conn.close()

@app.route("/")
def dashboard():
    conn = get_db()

    # Recent workouts
    recent_workouts = conn.execute(
        "SELECT * FROM workouts ORDER BY date DESC, id DESC LIMIT 10"
    ).fetchall()

    # Latest body stats
    latest_stats = conn.execute(
        "SELECT * FROM body_stats ORDER BY date DESC LIMIT 1"
    ).fetchone()

    # Body weight trend (last 30 days)
    weight_trend = conn.execute(
        "SELECT date, weight_kg FROM body_stats ORDER BY date ASC"
    ).fetchall()

    # Workout volume per exercise (for PRs)
    prs = conn.execute("""
        SELECT exercise, MAX(weight_kg) as max_weight, MAX(reps) as max_reps
        FROM workouts
        WHERE weight_kg > 0
        GROUP BY exercise
        ORDER BY exercise
    """).fetchall()

    # Workout count this week
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    workouts_this_week = conn.execute(
        "SELECT COUNT(DISTINCT date) FROM workouts WHERE date >= ?", (week_ago,)
    ).fetchone()[0]

    # Exercise frequency
    exercise_freq = conn.execute("""
        SELECT exercise, COUNT(*) as count
        FROM workouts
        GROUP BY exercise
        ORDER BY count DESC
        LIMIT 5
    """).fetchall()

    conn.close()

    return render_template("dashboard.html",
        recent_workouts=recent_workouts,
        latest_stats=latest_stats,
        weight_trend=weight_trend,
        prs=prs,
        workouts_this_week=workouts_this_week,
        exercise_freq=exercise_freq
    )

@app.route("/log-workout", methods=["GET", "POST"])
def log_workout():
    if request.method == "POST":
        conn = get_db()
        conn.execute(
            "INSERT INTO workouts (date, exercise, sets, reps, weight_kg, notes) VALUES (?,?,?,?,?,?)",
            (
                request.form["date"],
                request.form["exercise"],
                request.form.get("sets") or None,
                request.form.get("reps") or None,
                request.form.get("weight_kg") or 0,
                request.form.get("notes", "")
            )
        )
        conn.commit()
        conn.close()
        return redirect(url_for("dashboard"))
    return render_template("log_workout.html", today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/log-stats", methods=["GET", "POST"])
def log_stats():
    if request.method == "POST":
        conn = get_db()
        conn.execute(
            "INSERT INTO body_stats (date, weight_kg, body_fat_pct, notes) VALUES (?,?,?,?)",
            (
                request.form["date"],
                request.form.get("weight_kg") or None,
                request.form.get("body_fat_pct") or None,
                request.form.get("notes", "")
            )
        )
        conn.commit()
        conn.close()
        return redirect(url_for("dashboard"))
    return render_template("log_stats.html", today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/api/chart-data")
def chart_data():
    conn = get_db()
    weight_data = conn.execute(
        "SELECT date, weight_kg FROM body_stats ORDER BY date ASC"
    ).fetchall()

    exercise = request.args.get("exercise", "Bench Press")
    strength_data = conn.execute(
        "SELECT date, weight_kg FROM workouts WHERE exercise = ? ORDER BY date ASC",
        (exercise,)
    ).fetchall()

    exercises = conn.execute(
        "SELECT DISTINCT exercise FROM workouts ORDER BY exercise"
    ).fetchall()
    conn.close()

    return jsonify({
        "weight": {"labels": [r["date"] for r in weight_data], "values": [r["weight_kg"] for r in weight_data]},
        "strength": {"labels": [r["date"] for r in strength_data], "values": [r["weight_kg"] for r in strength_data]},
        "exercises": [r["exercise"] for r in exercises]
    })

def compute_goals(conn):
    """Attach current value + progress % to each goal."""
    goals = conn.execute("SELECT * FROM goals ORDER BY id DESC").fetchall()
    enriched = []
    for g in goals:
        g = dict(g)
        if g["type"] == "strength":
            row = conn.execute(
                "SELECT MAX(weight_kg) as cur FROM workouts WHERE exercise = ?",
                (g["exercise"],)
            ).fetchone()
            current = row["cur"] or g["start_kg"] or 0
        else:
            row = conn.execute(
                "SELECT weight_kg as cur FROM body_stats ORDER BY date DESC LIMIT 1"
            ).fetchone()
            current = row["cur"] if row else (g["start_kg"] or 0)

        g["current_kg"] = current
        start = g["start_kg"] or 0
        target = g["target_kg"]

        if g["type"] == "bodyweight":
            # lower is better
            if start <= target:
                pct = 100
            else:
                pct = max(0, min(100, round((start - current) / (start - target) * 100)))
        else:
            if start >= target:
                pct = 100
            else:
                pct = max(0, min(100, round((current - start) / (target - start) * 100)))

        g["progress_pct"] = pct
        g["done"] = pct >= 100

        # Days remaining
        if g["deadline"]:
            delta = (datetime.strptime(g["deadline"], "%Y-%m-%d") - datetime.now()).days
            g["days_left"] = delta
        else:
            g["days_left"] = None

        enriched.append(g)
    return enriched


@app.route("/goals", methods=["GET", "POST"])
def goals():
    conn = get_db()

    if request.method == "POST":
        goal_type = request.form["type"]
        exercise = request.form.get("exercise") or None
        target_kg = float(request.form["target_kg"])
        deadline = request.form.get("deadline") or None

        # Auto-detect start value
        if goal_type == "strength" and exercise:
            row = conn.execute(
                "SELECT MAX(weight_kg) as cur FROM workouts WHERE exercise = ?", (exercise,)
            ).fetchone()
            start_kg = row["cur"] or 0
        else:
            row = conn.execute(
                "SELECT weight_kg FROM body_stats ORDER BY date DESC LIMIT 1"
            ).fetchone()
            start_kg = row["weight_kg"] if row else 0

        conn.execute(
            "INSERT INTO goals (type, exercise, target_kg, start_kg, deadline) VALUES (?,?,?,?,?)",
            (goal_type, exercise, target_kg, start_kg, deadline)
        )
        conn.commit()
        conn.close()
        return redirect(url_for("goals"))

    all_goals = compute_goals(conn)
    exercises = [r["exercise"] for r in conn.execute(
        "SELECT DISTINCT exercise FROM workouts ORDER BY exercise"
    ).fetchall()]
    conn.close()

    return render_template("goals.html",
        goals=all_goals,
        exercises=exercises,
        today=datetime.now().strftime("%Y-%m-%d")
    )


@app.route("/goals/delete/<int:goal_id>", methods=["POST"])
def delete_goal(goal_id):
    conn = get_db()
    conn.execute("DELETE FROM goals WHERE id = ?", (goal_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("goals"))


@app.route("/history")
def history():
    conn = get_db()
    date_from = request.args.get("from", "")
    date_to   = request.args.get("to", "")
    exercise  = request.args.get("exercise", "")

    query  = "SELECT * FROM workouts WHERE 1=1"
    params = []
    if date_from:
        query += " AND date >= ?"; params.append(date_from)
    if date_to:
        query += " AND date <= ?"; params.append(date_to)
    if exercise:
        query += " AND exercise = ?"; params.append(exercise)
    query += " ORDER BY date DESC, id DESC"

    workouts  = conn.execute(query, params).fetchall()
    exercises = [r["exercise"] for r in conn.execute(
        "SELECT DISTINCT exercise FROM workouts ORDER BY exercise"
    ).fetchall()]
    conn.close()

    return render_template("history.html",
        workouts=workouts, exercises=exercises,
        date_from=date_from, date_to=date_to,
        selected_exercise=exercise,
        today=datetime.now().strftime("%Y-%m-%d")
    )


@app.route("/weekly")
def weekly():
    conn = get_db()
    today      = datetime.now()
    week_start = today - timedelta(days=today.weekday())   # Monday
    prev_start = week_start - timedelta(days=7)

    def week_range(start):
        return start.strftime("%Y-%m-%d"), (start + timedelta(days=6)).strftime("%Y-%m-%d")

    cur_from,  cur_to  = week_range(week_start)
    prev_from, prev_to = week_range(prev_start)

    def week_stats(d_from, d_to):
        sessions = conn.execute(
            "SELECT COUNT(DISTINCT date) FROM workouts WHERE date BETWEEN ? AND ?",
            (d_from, d_to)
        ).fetchone()[0]

        total_sets = conn.execute(
            "SELECT COALESCE(SUM(sets),0) FROM workouts WHERE date BETWEEN ? AND ?",
            (d_from, d_to)
        ).fetchone()[0]

        total_volume = conn.execute(
            """SELECT COALESCE(SUM(sets * reps * weight_kg), 0)
               FROM workouts WHERE date BETWEEN ? AND ? AND weight_kg > 0""",
            (d_from, d_to)
        ).fetchone()[0]

        prs = conn.execute(
            """SELECT w.exercise, w.weight_kg, w.date
               FROM workouts w
               WHERE w.date BETWEEN ? AND ? AND w.weight_kg > 0
               AND w.weight_kg = (
                   SELECT MAX(w2.weight_kg) FROM workouts w2
                   WHERE w2.exercise = w.exercise AND w2.date <= ?
               )
               AND NOT EXISTS (
                   SELECT 1 FROM workouts w3
                   WHERE w3.exercise = w.exercise
                   AND w3.weight_kg >= w.weight_kg
                   AND w3.date < ?
               )""",
            (d_from, d_to, d_to, d_from)
        ).fetchall()

        exercises = conn.execute(
            """SELECT exercise, COUNT(*) as cnt, MAX(weight_kg) as top_weight
               FROM workouts WHERE date BETWEEN ? AND ?
               GROUP BY exercise ORDER BY cnt DESC""",
            (d_from, d_to)
        ).fetchall()

        body = conn.execute(
            "SELECT weight_kg FROM body_stats WHERE date BETWEEN ? AND ? ORDER BY date DESC LIMIT 1",
            (d_from, d_to)
        ).fetchone()

        return dict(
            sessions=sessions, total_sets=total_sets,
            total_volume=round(total_volume),
            prs=prs, exercises=exercises,
            body_weight=body["weight_kg"] if body else None
        )

    cur  = week_stats(cur_from,  cur_to)
    prev = week_stats(prev_from, prev_to)

    def delta(a, b):
        if b == 0: return None
        return round(((a - b) / b) * 100, 1)

    comparison = dict(
        sessions = delta(cur["sessions"],     prev["sessions"]),
        volume   = delta(cur["total_volume"], prev["total_volume"]),
        sets     = delta(cur["total_sets"],   prev["total_sets"]),
    )

    conn.close()
    return render_template("weekly.html",
        cur=cur, prev=prev, comparison=comparison,
        cur_from=cur_from, cur_to=cur_to,
        prev_from=prev_from, prev_to=prev_to
    )


@app.route("/export")
def export_excel():
    conn = get_db()
    workouts   = conn.execute("SELECT * FROM workouts ORDER BY date DESC").fetchall()
    body_stats = conn.execute("SELECT * FROM body_stats ORDER BY date DESC").fetchall()
    goals_raw  = conn.execute("SELECT * FROM goals ORDER BY id").fetchall()
    goals_full = compute_goals(conn)
    conn.close()

    wb = openpyxl.Workbook()

    # ── Colour palette ──────────────────────────────────────────────
    BLACK  = "0A0A0A"
    YELLOW = "E8FF47"
    ORANGE = "FF6B35"
    DGREY  = "1A1A1A"
    MGREY  = "2A2A2A"
    WHITE  = "F0F0F0"
    GREEN  = "4ADE80"

    def hdr_style(cell, bg=BLACK, fg=YELLOW, bold=True, size=11):
        cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_style(cell, bg=DGREY, fg=WHITE, bold=False, align="left"):
        cell.font      = Font(color=fg, size=10, bold=bold, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")

    def thin_border():
        s = Side(style="thin", color="333333")
        return Border(left=s, right=s, top=s, bottom=s)

    def write_table(ws, headers, rows, col_widths=None):
        # Header row
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            hdr_style(cell)
            cell.border = thin_border()
        ws.row_dimensions[1].height = 22

        # Data rows
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                bg = DGREY if r % 2 == 0 else MGREY
                data_style(cell, bg=bg)
                cell.border = thin_border()
            ws.row_dimensions[r].height = 18

        # Column widths
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ── Sheet 1 : Workouts ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Workouts"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = YELLOW

    workout_rows = [
        (w["date"], w["exercise"], w["sets"], w["reps"],
         w["weight_kg"], round((w["sets"] or 0) * (w["reps"] or 0) * (w["weight_kg"] or 0), 1),
         w["notes"] or "")
        for w in workouts
    ]
    write_table(ws1,
        ["Date", "Exercise", "Sets", "Reps", "Weight (kg)", "Volume (kg)", "Notes"],
        workout_rows,
        [14, 20, 8, 8, 13, 13, 30]
    )

    # ── Sheet 2 : Body Stats ─────────────────────────────────────────
    ws2 = wb.create_sheet("Body Stats")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = ORANGE

    stat_rows = [
        (s["date"], s["weight_kg"], s["body_fat_pct"], s["notes"] or "")
        for s in body_stats
    ]
    write_table(ws2,
        ["Date", "Weight (kg)", "Body Fat (%)", "Notes"],
        stat_rows,
        [14, 14, 14, 30]
    )

    # ── Sheet 3 : Goals ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Goals")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = GREEN

    goal_rows = []
    for g in goals_full:
        status = "✓ Achieved" if g["done"] else f"{g['progress_pct']}%"
        goal_rows.append((
            g["type"].title(),
            g["exercise"] or "—",
            g["start_kg"],
            g["current_kg"],
            g["target_kg"],
            status,
            g["deadline"] or "—",
            g["days_left"] if g["days_left"] is not None else "—"
        ))
    write_table(ws3,
        ["Type", "Exercise", "Start (kg)", "Current (kg)", "Target (kg)", "Progress", "Deadline", "Days Left"],
        goal_rows,
        [14, 16, 13, 14, 13, 12, 14, 11]
    )

    # Colour the progress cell green if achieved
    for r, g in enumerate(goals_full, 2):
        cell = ws3.cell(row=r, column=6)
        if g["done"]:
            cell.font = Font(color=GREEN, bold=True, size=10, name="Calibri")
        else:
            cell.font = Font(color=YELLOW, bold=True, size=10, name="Calibri")

    # ── Sheet 4 : Summary ────────────────────────────────────────────
    ws4 = wb.create_sheet("Summary")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "888888"

    def s(row, col, value, bg=BLACK, fg=WHITE, bold=False, size=11, align="left"):
        cell = ws4.cell(row=row, column=col, value=value)
        cell.font      = Font(color=fg, bold=bold, size=size, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        return cell

    # Title
    ws4.merge_cells("A1:D1")
    t = ws4["A1"]
    t.value = "GYMTRACK — EXPORT SUMMARY"
    t.font  = Font(bold=True, color=YELLOW, size=14, name="Calibri")
    t.fill  = PatternFill("solid", fgColor=BLACK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    s(2, 1, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", fg="888888", size=9)

    stats_data = [
        ("Total Workouts Logged", len(workouts)),
        ("Unique Exercises",      len(set(w["exercise"] for w in workouts))),
        ("Body Stat Entries",     len(body_stats)),
        ("Active Goals",          len([g for g in goals_full if not g["done"]])),
        ("Goals Achieved",        len([g for g in goals_full if g["done"]])),
    ]
    for i, (label, val) in enumerate(stats_data, 4):
        s(i, 1, label, bg=DGREY, fg=WHITE, size=11)
        s(i, 2, val,   bg=DGREY, fg=YELLOW, bold=True, size=13, align="center")
        ws4.row_dimensions[i].height = 22

    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 16
    ws4.column_dimensions["D"].width = 16

    # ── Stream to browser ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"gymtrack_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    app.run(debug=True)
